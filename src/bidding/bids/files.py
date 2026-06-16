# This file contains classes to access Excel files
import openpyxl
import csv
from pathlib import Path
from utils import Asset, MyFileAccessException


class SemiColon(csv.Dialect):
   # Describes the wished usual properties for CSV files.
    delimiter = ';'
    quotechar = '"'
    doublequote = True
    skipinitialspace = True
    lineterminator = '\r\n'
    quoting = csv.QUOTE_MINIMAL


class CsvFile():
   def __init__(self, file_name: Path, fields: list[str]):
      self.fields = fields
      self.name = file_name

   def exists(self) -> bool:
      return self.name.exists()
   
   def recreate(self):
      # This function erase existing file if any and creates empty one with header.
      try:
         with open(self.name, "w", newline='') as csvfile:
            writer = csv.DictWriter(csvfile, dialect=SemiColon, fieldnames=self.fields)
            writer.writeheader()
      except csv.Error as error:
         print(error)
         raise MyFileAccessException(self.name, *error.__str__())
   
   def add_row(self, instance_dict):
      # This function adds a new row with given data at the end of file.
      try:
         with open(self.name, "a+", newline='') as csvfile:
            writer = csv.DictWriter(csvfile, dialect=SemiColon, fieldnames=self.fields)
            writer.writerow(instance_dict)
      except Exception as error:
         raise MyFileAccessException(self.name, *error.__str__())

   def get_row(self, id: int) -> dict:
      # This function returns row for given id, or empty if no row found.
      with open(self.name, newline='') as csvfile:
         rows = csv.DictReader(csvfile, dialect=SemiColon)
         for row in rows:
            if int(row['id']) == id:
               return row
   
   def get_last_row(self) -> dict[str]:
      # This function returns last row of file.
      with open(self.name, newline='') as csvfile:
         reader = reversed(list(csv.DictReader(csvfile, dialect=SemiColon)))
         last_row = self._process_iterator(reader)
         csvfile.close()
         return last_row
             
   def _process_iterator(self, iterator, default=None):
      # This function returns next item in given iterator or None if no next
      try:
         return next(iterator)
      except StopIteration:
         return default


class RuleFile(CsvFile):
   def __init__(self, fields: list[str]):
      name = Asset.path("bid_rules.csv")
      super().__init__(name, fields)

   def get_rows(self, step: str = "") -> list:
      with open(self.name, newline='') as csvfile:
         matching_rows = []
         rows = csv.DictReader(csvfile, dialect=SemiColon)
         for row in rows:
            if row['step'] == step or not step:
               matching_rows.append(row)      
      return matching_rows


class SenseFile(CsvFile):
   def __init__(self, fields: list[str]):
      name = Asset.path("bid_senses.csv")
      super().__init__(name, fields)


class BidHistoryFile(CsvFile):
   def __init__(self, fields: list[str]):
      name = Asset.path("bid_history.csv")
      super().__init__(name, fields)

   def get_rule_rows(self) -> list:
      with open(self.name, newline='') as csvfile:
         matching_rows = []
         rows = csv.DictReader(csvfile, dialect=SemiColon)
         for row in rows:
            if row['type'] == "RULE":
               matching_rows.append(row)      
      return matching_rows



class ExcelFile:
   def __init__(self, wb: openpyxl.Workbook, sheet_name: str):
      self.sheet = wb[sheet_name]
      self.header = self._get_header(self.sheet)
      self.column_range = range(0, len(self.header))

   def _get_header(self, excel_sheet) -> list:
      # Return list of relevant fields that are in row 0.
      fields = []
      for col in excel_sheet.iter_cols(0, excel_sheet.max_column):
         title = str(col[0].value)
         if title == "SEF_page":
            break
         else:
            fields.append(title)
      return fields


class RuleExcelFile(ExcelFile):
   def __init__(self):
      file_name = Asset.path("SEFrule.xlsx")
      wb = openpyxl.load_workbook(file_name)
      super().__init__(wb=wb, sheet_name="rules")


class SenseExcelFile(ExcelFile):
   def __init__(self):
      file_name = Asset.path("SEFsense.xlsx")
      wb = openpyxl.load_workbook(file_name)
      super().__init__(wb=wb, sheet_name="bids")


class ExcelToCsv:
   ALLOW_BLANK = [
      "distribution",
      "hist_bid",
      "convention",
      "comment",
   ]
   BOOL_FIELDS = [
      "suit_stop",
      "suit_control",
      "suit_force",
      "opp_stop",
      "artificial",
      "awake",
   ]
   NUM_OP_FIELDS = [
      "distribution",
      "spade_count",
      "heart_count",
      "diamond_count",
      "club_count",
      "par_suit_count",
      "suit_count",
      "suit1_count",
      "suit2_count",
      "first_pass",
      "fit_cards",
      "arg2",
      "arg_bid",
   ]
   NUMERIC_FIELDS = [
      "won_tricks",
      "lost_tricks",
      "stops",
      "sense_id",
   ]

   def convert_rules(self, csv_fields: list[str]) -> RuleFile:
      excel_file = RuleExcelFile()
      csv_file = RuleFile(csv_fields)
      return self._convert(excel_file, csv_file)

   def convert_senses(self, csv_fields: list[str]) -> SenseFile:
      excel_file = SenseExcelFile()
      csv_file = SenseFile(csv_fields)
      return self._convert(excel_file, csv_file)

   def _convert(self, excel_file: ExcelFile, csv_file: CsvFile):
      try:
         csv_file.recreate()
         self._write_csv_row(excel_file, csv_file)
      except Exception as error:
         raise error
      return csv_file

   def _write_csv_row(self, excel_file: ExcelFile, csv_file: CsvFile):
      # Writes a row in csv file for each row read in excel sheet.
      sheet = excel_file.sheet
      with open(csv_file.name, "a+", newline='') as csvfile:
         writer = csv.DictWriter(csvfile, dialect=SemiColon, fieldnames=csv_file.fields)
         try:
            for excel_row in sheet.iter_rows(min_row=2, values_only=True):
               row = {}
               for field in csv_file.fields:
                  col = excel_file.header.index(field)
                  row[field] = self._normalize(field, excel_row[col])
               writer.writerow(row)
         except Exception as error:
            raise error
   
   def _normalize(self, field: str, excel_value) -> any:
      # This function adapts default values and formats.
      if isinstance(excel_value, str) and not field in self.ALLOW_BLANK:
            excel_value = excel_value.replace(" ", "")
            
      if field == "hist_bid" and excel_value:
         return excel_value.replace("-", "passe")      
      elif field in self.BOOL_FIELDS:
         return True if excel_value == 1 else False
      elif field in self.NUM_OP_FIELDS:
         return str(excel_value) if excel_value else ""
      elif field in self.NUMERIC_FIELDS:
         return excel_value if excel_value else 0
      else:
         return excel_value if excel_value else ""
