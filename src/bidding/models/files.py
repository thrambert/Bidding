# This file contains classes to access Excel files
import openpyxl
import csv
from utils import Asset, MyFileAccessException


class SemiColon(csv.Dialect):
   # Describes the wished usual properties for CSV files.
    delimiter = ';'
    quotechar = '"'
    doublequote = True
    skipinitialspace = True
    lineterminator = '\r\n'
    quoting = csv.QUOTE_MINIMAL


class BidRuleFile():
   def __init__(self, fields: list[str]):
      self.fields = fields
      self.name = Asset.path("bidtree.csv")

   def recreate(self):
      # This function erase existing file if any and creates empty one with header.
      try:
         with open(self.name, "w", newline='') as csvfile:
            writer = csv.DictWriter(csvfile, dialect=SemiColon, fieldnames=self.fields)
            writer.writeheader()
      except csv.Error as error:
         print(error)
         raise MyFileAccessException(self.name, *error.__str__())
   
   def get_row(self, id: int) -> dict:
      # This function returns row for given id, or empty if no row found.
      with open(self.name, newline='') as csvfile:
         rows = csv.DictReader(csvfile, dialect=SemiColon)
         for row in rows:
            if int(row['id']) == id:
               return row
  
   def get_rows(self, situation: str, sequence: str) -> list:
      # This function returns a list of rows matching given arguments.
      with open(self.name, newline='') as csvfile:
         matching_rows = []
         rows = csv.DictReader(csvfile, dialect=SemiColon)
         for row in rows:
            if row['situation'] == situation and row['sequence'] == sequence:
               matching_rows.append(row)      
      return matching_rows


class BidExcelFile:
   def __init__(self):
      file_name = Asset.path("bidtree.xlsx")
      wb = openpyxl.load_workbook(file_name)
      self.sheet = wb["rules"]
      self.header = self._get_header(self.sheet)
      self.column_range = range(0, len(self.header) - 1)

   def _get_header(self, excel_sheet) -> list:
      # Return list of relevant fields that are in row 0.
      fields = []
      for col in excel_sheet.iter_cols(0, excel_sheet.max_column):
         title = str(col[0].value)
         if title == "Remarque":
            break
         else:
            fields.append(title)
      return fields


class BidFileConverter:
   def excel_to_Csv(self, csv_fields: list[str]) -> BidRuleFile:
      # This function creates csv file from excel file.
      excel_file = BidExcelFile()
      csv_file = BidRuleFile(csv_fields)
      try:
         csv_file.recreate()
         self._write_bid_tree_csv(excel_file, csv_file)
      except Exception as error:
         raise error
      return csv_file

   def _write_bid_tree_csv(self, excel: BidExcelFile, csv_file: BidRuleFile):
      # Writes a row in csv file for each row read in excel sheet.
      with open(csv_file.name, "a+", newline='') as csvfile:
         writer = csv.DictWriter(csvfile, dialect=SemiColon, fieldnames=csv_file.fields)
         try:
            for excel_row in excel.sheet.iter_rows(min_row=3, max_row=excel.sheet.max_row, values_only=True):
               row = {}
               for col in excel.column_range:
                  row[excel.header[col]] = self._normalize(excel.header[col], excel_row[col], col)
               writer.writerow(row)
         except Exception as error:
            raise error
   
   def _normalize(self, field: str, excel_value, col: int) -> any:
      # This function adapts default values and formats.
      if isinstance(excel_value, str) and (field == "points" or col in range(8, 25)):
         excel_value = excel_value.replace(" ", "")
         
      if field in ["awake", "artificial"]:
         return True if excel_value == 1 else False
      elif field in ["distribution", "color1_count", "color2_count", "fit_cards"]:
         return str(excel_value) if excel_value else ""
      elif field in ["SEF_page", "won_tricks", "lost_tricks"]:
         return excel_value if excel_value else 0
      else:
         return excel_value if excel_value else ""
